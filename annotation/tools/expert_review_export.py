#!/usr/bin/env python3
"""
Export annotations for expert oncologist review.

Generates a simple Excel file where medical experts can:
- Read the abstract
- See the extracted entities (listed plainly)
- See the proposed relations in plain language
- Mark each relation as: Correct / Incorrect / Unsure
- Add comments

Designed to be as simple as possible for non-technical reviewers.

Usage:
    python expert_review_export.py --input ../data/validated.jsonl \
                                   --output ../data/expert_review.xlsx \
                                   --n 50
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


ENTITY_SHORT = {
    "GeneOrProtein": "Gene/Protein",
    "Chemical": "Drug/Chemical",
    "Disease": "Disease",
    "SequenceVariant": "Mutation/Variant",
    "Pathway": "Pathway",
    "CellLine": "Cell Line"
}

RELATION_PLAIN = {
    "Positive_Correlation": "increases / activates / causes",
    "Negative_Correlation": "decreases / inhibits / treats",
    "Association": "is associated with",
    "Binding": "physically binds to",
    "Drug_Interaction": "interacts with (pharmacologically)",
    "Cotreatment": "is co-administered with",
    "Comparison": "is compared with",
    "Conversion": "is converted to"
}


def format_relation_plain(rel: dict, ent_lookup: dict) -> str:
    """Convert a relation to plain English for expert review."""
    subj = ent_lookup.get(rel.get("subject_id", ""), {})
    obj = ent_lookup.get(rel.get("object_id", ""), {})

    subj_text = subj.get("text", "?")
    subj_type = ENTITY_SHORT.get(subj.get("type", ""), "?")
    obj_text = obj.get("text", "?")
    obj_type = ENTITY_SHORT.get(obj.get("type", ""), "?")

    rel_type = rel.get("type", "?")
    rel_plain = RELATION_PLAIN.get(rel_type, rel_type)
    novelty = rel.get("novelty", "?")

    return f"{subj_text} [{subj_type}]  {rel_plain}  {obj_text} [{obj_type}]  ({novelty})"


def build_expert_workbook(records: list) -> Workbook:
    """Build an Excel workbook for expert review."""
    wb = Workbook()

    # --- Instructions sheet ---
    ws_instr = wb.active
    ws_instr.title = "Instructions"

    header_font = Font(name="Calibri", size=14, bold=True)
    body_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)

    ws_instr["A1"] = "PanCan-RE Expert Review"
    ws_instr["A1"].font = header_font

    instructions = [
        "",
        "Thank you for reviewing these annotations. Your expert validation is essential.",
        "",
        "HOW TO REVIEW:",
        "1. Go to the 'Review' sheet",
        "2. For each row, read the Abstract and the Proposed Relation",
        "3. In the 'Your Verdict' column, enter one of:",
        "      Correct  —  the relation is accurate",
        "      Incorrect  —  the relation is wrong or not supported by the text",
        "      Unsure  —  you are not certain",
        "4. Optionally add a comment explaining your reasoning",
        "",
        "RELATION TYPES:",
        "  • increases / activates / causes  —  Entity A promotes or causes Entity B",
        "  • decreases / inhibits / treats  —  Entity A reduces or blocks Entity B",
        "  • is associated with  —  Entities are linked but direction is unclear",
        "  • physically binds to  —  Direct molecular binding",
        "  • interacts with  —  Pharmacological drug interaction",
        "  • is co-administered with  —  Drugs given together as treatment",
        "  • is compared with  —  Entities compared in the study",
        "  • is converted to  —  Chemical metabolism/conversion",
        "",
        "NOVELTY:",
        "  • Novel  —  This is a new finding of the study",
        "  • Known  —  This is established background knowledge",
    ]

    for i, line in enumerate(instructions, start=2):
        ws_instr[f"A{i}"] = line
        ws_instr[f"A{i}"].font = bold_font if line.startswith(("HOW", "RELATION", "NOVELTY")) else body_font

    ws_instr.column_dimensions["A"].width = 80

    # --- Review sheet ---
    ws = wb.create_sheet("Review")

    # Headers
    headers = [
        "Record ID",
        "Title",
        "Abstract",
        "Entities Found",
        "Proposed Relation",
        "Your Verdict\n(Correct / Incorrect / Unsure)",
        "Your Comment"
    ]

    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font_style = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Column widths
    widths = [14, 40, 60, 35, 50, 18, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows — one row per RELATION (not per record)
    row_num = 2
    verdict_fill_even = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    verdict_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for rec in records:
        entities = rec.get("entities", [])
        relations = rec.get("relations", [])
        ent_lookup = {e["entity_id"]: e for e in entities}

        # Format entities list
        ent_list = "\n".join(
            f"• {e.get('text', '?')} [{ENTITY_SHORT.get(e.get('type', ''), '?')}]"
            for e in entities
        )

        if not relations:
            continue

        for rel in relations:
            rel_plain = format_relation_plain(rel, ent_lookup)
            bg = verdict_fill_even if row_num % 2 == 0 else verdict_fill_odd

            data = [
                rec.get("record_id", ""),
                rec.get("title", ""),
                rec.get("abstract", "")[:1000],  # Truncate for readability
                ent_list,
                rel_plain,
                "",  # Verdict — expert fills this
                ""   # Comment — expert fills this
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if col_idx in (6, 7):
                    cell.fill = bg

            ws.row_dimensions[row_num].height = 80
            row_num += 1

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add data validation for verdict column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Correct,Incorrect,Unsure"',
        allow_blank=True,
        showDropDown=False
    )
    dv.prompt = "Select your verdict"
    dv.promptTitle = "Verdict"
    ws.add_data_validation(dv)
    dv.add(f"F2:F{row_num - 1}")

    # --- Summary sheet ---
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "PanCan-RE Expert Review Summary"
    ws_summary["A1"].font = header_font
    ws_summary["A3"] = f"Records: {len(records)}"
    ws_summary["A4"] = f"Total relations to review: {row_num - 2}"
    ws_summary["A5"] = f"Generated: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary["A7"] = "After completing your review, please return this file."
    ws_summary["A7"].font = bold_font
    ws_summary.column_dimensions["A"].width = 60

    return wb


def main():
    parser = argparse.ArgumentParser(description="Export annotations for expert review")
    parser.add_argument("--input", default="../data/validated.jsonl")
    parser.add_argument("--output", default="../data/expert_review.xlsx")
    parser.add_argument("--n", type=int, default=50, help="Number of records to include")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    input_path = str((script_dir / args.input).resolve())
    output_path = str((script_dir / args.output).resolve())

    print(f"Loading validated annotations from {input_path}...")
    records = []
    with open(input_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rec = json.loads(line)
                annotator = rec.get("metadata", {}).get("annotator", "")
                if annotator and annotator != "groq_preannotation":
                    records.append(rec)

    if not records:
        print("ERROR: No validated records found.")
        sys.exit(1)

    # Take subset
    records = records[:args.n]
    print(f"Exporting {len(records)} records for expert review...")

    wb = build_expert_workbook(records)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    total_relations = sum(len(r.get("relations", [])) for r in records)
    print(f"\nExpert review file saved: {output_path}")
    print(f"  Records: {len(records)}")
    print(f"  Relations to review: {total_relations}")
    print(f"\nSend this Excel file to the oncology experts.")


if __name__ == "__main__":
    main()
